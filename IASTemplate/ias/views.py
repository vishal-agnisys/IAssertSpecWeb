import os
import subprocess
from subprocess import *
from subprocess import SubprocessError

from openpyxl import load_workbook
from django.shortcuts import render

from .models import Post


def index(request):
    return render(request, 'ias/index.html', {})


def input(request):
    if request.method == 'POST':

        return render(request, 'ias/input.html', {})
    else:
        return render(request, 'ias/input.html', {})


def output(request):
    if request.method == 'POST':
        input_data = request.POST.get('input_field')
        input_data = str(input_data)
        email_id = request.POST.get('email_id')
        rows = counter()
        row = rows
        if email_id == '' or email_id == 'None':
            email_id = 'Returning_user'
            insert_email_Id(email_id, row)
            # print("returning User")
        else:
            insert_email_Id(email_id, row)
            # print("Email-id of user is  " + str(email_id))
        if input_data == 'None':
            input_data = 'assert property (@(negedge clock) A |-> (B & C));'
            output_data = ''
        else:
            input_data = str(input_data)
            output_data = getAssertion(input_data)
            insert_data(input_data, output_data, row)
        # print("input data   :" + input_data)

        sugest_data = request.POST.get('sugest_out_field')
        if sugest_data == 'None' or sugest_data == '':
            sugest_data = 'Assertion as Expected'
            insert_suggest_output(sugest_data, row)
            # print("correct assertion")
        else:
            insert_suggest_output(sugest_data, row)
            # print("suggested output is:=  " + str(sugest_data))

        response_data = {}
        post = Post(input_field=input_data, output_field=output_data)
        response_data['input_field'] = request.POST.get('input_field')
        response_data['output_field'] = post.output_field

        return render(request, 'ias/output.html', {'response_data': post})
    else:
        return render(request, 'ias/input.html', {})


def jarWrapper(args):
    arg = ""
    # print("Returning arg:= " + str(arg))
    process = subprocess.Popen(['java', '-jar', 'IAssertSpec.jar'] + args, stdout=PIPE, stderr=PIPE)
    # print("Returning arg Value is: " + str(args[0]))
    stdout, stderr = process.communicate()
    # print("final result   " + str(stdout, 'utf-8'))
    stdout = str(stdout, 'utf-8')
    stdout = listToStringWithoutBrackets(stdout)
    arg = str(stdout)
    # print("result value:=   " + arg)
    return arg


def listToStringWithoutBrackets(list1):
    return str(list1).replace('[','').replace(']','')


os.system('cls')


def getAssertion(input_field_data):
    arg1 = [input_field_data]
    result = jarWrapper(arg1)
    return result


def insert_data(arg1, arg2, row):
    filepath = "IAssertSpecData.xlsx"
    row = row
    col = 2
    wb = load_workbook(filepath)
    sheet = wb.active
    sheet.cell(row=row, column=col).value = arg1
    sheet.cell(row=row, column=col + 1).value = arg2
    wb.save(filepath)


def insert_suggest_output(arg1, row):
    filepath = "IAssertSpecData.xlsx"
    row = row
    col = 4
    wb = load_workbook(filepath)
    sheet = wb.active
    sheet.cell(row=row-1, column=col).value = arg1
    wb.save(filepath)


def insert_email_Id(arg1, row):
    filepath = "IAssertSpecData.xlsx"
    row = row
    col = 1
    wb = load_workbook(filepath)
    sheet = wb.active
    sheet.cell(row=row, column=col).value = arg1
    wb.save(filepath)


def counter():
    if 'cnt' not in counter.__dict__:
        counter.cnt = 1
    counter.cnt += 1
    # print("counter value   " + str(counter.cnt))
    return counter.cnt



